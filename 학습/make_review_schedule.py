from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Git 커밋 기준 프로그래머스 Lv.2~3 문제 (중복 제거, 날짜 내림차순)
problems = [
    ("2026-06-24", "피로도",              "Lv.2", "완전탐색"),
    ("2026-06-24", "소수 찾기",           "Lv.2", "완전탐색"),
    ("2026-06-22", "지게차와 크레인",     "Lv.2", "구현"),
    ("2026-06-21", "섬 연결하기",         "Lv.3", "그리디"),
    ("2026-06-20", "모음 사전",           "Lv.2", "완전탐색"),
    ("2026-06-20", "전력망을 둘로 나누기","Lv.2", "BFS/DFS"),
    ("2026-06-20", "게임 맵 최단거리",   "Lv.2", "BFS/DFS"),
    ("2026-06-19", "혼자 놀기의 달인",   "Lv.2", "그래프"),
    ("2026-06-19", "영어 끝말잇기",      "Lv.2", "구현"),
    ("2026-06-11", "퍼즐 조각 채우기",   "Lv.3", "BFS/DFS"),
    ("2026-06-10", "여행경로",           "Lv.3", "BFS/DFS"),
    ("2026-06-10", "이중우선순위큐",     "Lv.3", "힙"),
    ("2026-06-05", "조이스틱",           "Lv.2", "그리디"),
    ("2026-06-04", "구명보트",           "Lv.2", "그리디"),
    ("2026-06-04", "큰 수 만들기",       "Lv.2", "그리디"),
    ("2026-06-01", "미로 탈출",          "Lv.2", "BFS/DFS"),
    ("2026-05-28", "베스트앨범",         "Lv.3", "해시"),
    ("2026-05-24", "H-Index",            "Lv.2", "정렬"),
    ("2026-05-22", "주식가격",           "Lv.2", "스택/큐"),
    ("2026-05-22", "다리를 지나는 트럭", "Lv.2", "스택/큐"),
    ("2026-05-21", "입국심사",           "Lv.3", "이분탐색"),
    ("2026-05-21", "가장 큰 수",         "Lv.2", "정렬"),
    ("2026-05-20", "의상",               "Lv.2", "해시"),
    ("2026-05-19", "가장 먼 노드",       "Lv.3", "BFS/DFS"),
    ("2026-05-12", "전화번호 목록",      "Lv.2", "해시"),
    ("2026-04-04", "단어 변환",          "Lv.3", "BFS/DFS"),
    ("2026-04-03", "카펫",               "Lv.2", "완전탐색"),
    ("2026-04-02", "타겟 넘버",          "Lv.2", "BFS/DFS"),
    ("2026-04-02", "네트워크",           "Lv.3", "BFS/DFS"),
    ("2025-11-04", "길 찾기 게임",       "Lv.3", "트리"),
    ("2025-09-29", "더 맵게",            "Lv.2", "힙"),
    ("2025-09-23", "프로세스",           "Lv.2", "스택/큐"),
    ("2025-09-16", "올바른 괄호",        "Lv.2", "스택/큐"),
    ("2025-05-01", "예상 대진표",        "Lv.2", "수학"),
]

# ── 스타일 정의 ──────────────────────────────────────────
LEVEL_FILL = {
    "Lv.1": PatternFill("solid", fgColor="C6EFCE"),  # 초록
    "Lv.2": PatternFill("solid", fgColor="FFEB9C"),  # 노랑
    "Lv.3": PatternFill("solid", fgColor="FFCC99"),  # 주황
}
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
CHECK_FILL  = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 워크북 생성 ──────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "1-3-7 복습 스케줄"

headers    = ["#", "문제명", "레벨", "카테고리", "푼 날짜",
              "+1일 복습일", "+1일 완료",
              "+3일 복습일", "+3일 완료",
              "+7일 복습일", "+7일 완료",
              "메모 (틀린 이유 / 핵심 아이디어)"]
col_widths = [5, 24, 8, 12, 13, 13, 10, 13, 10, 13, 10, 42]

# 헤더 행
for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = CENTER
    cell.border    = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"  # 헤더 고정

# 데이터 행
for row_idx, (date_str, name, level, category) in enumerate(problems, start=2):
    solved = datetime.strptime(date_str, "%Y-%m-%d")
    d1 = (solved + timedelta(days=1)).strftime("%Y-%m-%d")
    d3 = (solved + timedelta(days=3)).strftime("%Y-%m-%d")
    d7 = (solved + timedelta(days=7)).strftime("%Y-%m-%d")

    row_data = [row_idx - 1, name, level, category, date_str,
                d1, "", d3, "", d7, "", ""]
    lv_fill = LEVEL_FILL[level]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.font   = BODY_FONT

        if col_idx in (7, 9, 11):   # 완료 체크 칸
            cell.fill      = CHECK_FILL
            cell.alignment = CENTER
        elif col_idx == 12:          # 메모 칸
            cell.fill      = WHITE_FILL
            cell.alignment = LEFT
        else:
            cell.fill      = lv_fill
            cell.alignment = CENTER

    ws.row_dimensions[row_idx].height = 22

# ── 범례 시트 ────────────────────────────────────────────
ws2 = wb.create_sheet("범례")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 30

legend = [
    ("색상",          "의미",          HEADER_FILL, HEADER_FONT),
    ("초록 (Lv.1)",   "난이도 낮음",   LEVEL_FILL["Lv.1"], BODY_FONT),
    ("노랑 (Lv.2)",   "난이도 중간",   LEVEL_FILL["Lv.2"], BODY_FONT),
    ("주황 (Lv.3)",   "난이도 높음",   LEVEL_FILL["Lv.3"], BODY_FONT),
    ("회색",          "+1/3/7일 완료 체크 칸", CHECK_FILL,  BODY_FONT),
]
for r, (a, b, fill, font) in enumerate(legend, start=1):
    for c, val in enumerate([a, b], start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = CENTER
        cell.border    = BORDER
    ws2.row_dimensions[r].height = 22

out_path = r"C:\Users\허민엽\Desktop\컴공\개인공부\프로젝트\ps\복습_스케줄.xlsx"
wb.save(out_path)
print(f"저장 완료: {out_path}")
print(f"총 문제 수: {len(problems)}개")
